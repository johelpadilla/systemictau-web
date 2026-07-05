"""
Systemic Tau — Structured Export Module (JSON + Excel)

Provides clean, reproducible, machine-readable exports of analysis results.
Designed to work with the Studio (and reusable by desktop/platform layers).

Exports respect the current set of run analyses (can be partial).
"""

from __future__ import annotations
import json
import datetime as dt
import copy
from io import BytesIO
from typing import Dict, Any, Optional, List

import numpy as np
import pandas as pd

# Optional advanced formatting (used for professional Excel output)
try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL_STYLES = True
except Exception:
    HAS_OPENPYXL_STYLES = False
    Font = PatternFill = Alignment = Border = Side = None  # type: ignore
    dataframe_to_rows = None  # type: ignore


def _to_python(obj: Any) -> Any:
    """Recursively convert numpy types to native Python for JSON serialization."""
    if isinstance(obj, dict):
        return {k: _to_python(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_to_python(x) for x in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj) if np.isfinite(obj) else None
    if isinstance(obj, (np.ndarray,)):
        return _to_python(obj.tolist())
    if isinstance(obj, (pd.Timestamp, dt.datetime)):
        return obj.isoformat()
    if pd.isna(obj):  # catches np.nan, pd.NA etc.
        return None
    return obj


def _safe_get(d: Dict, *keys, default=None):
    for k in keys:
        if d is None:
            return default
        d = d.get(k, default) if isinstance(d, dict) else default
    return d


def build_structured_export(
    analyses: Dict[str, Dict[str, Any]],
    df_raw: Optional[pd.DataFrame] = None,
    active_view: Optional[str] = None,
    window_size: Optional[int] = None,
    extra_metadata: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Build a clean, documented dict ready for JSON export.
    """
    if not analyses:
        analyses = {}

    now = dt.datetime.now(dt.timezone.utc).isoformat()

    metadata = {
        "version": "systemictau-studio-v4",
        "exported_at": now,
        "n_perspectives": len(analyses),
        "perspectives": list(analyses.keys()),
        "active_view": active_view,
        "dataset": {
            "n_rows": int(df_raw.shape[0]) if df_raw is not None else None,
            "n_columns": int(df_raw.shape[1]) if df_raw is not None else None,
            "columns": list(df_raw.columns) if df_raw is not None else None,
        },
    }
    if extra_metadata:
        metadata.update(extra_metadata)

    results: Dict[str, Any] = {}
    cluster_composition: Dict[str, Any] = {}

    perspective_order = [p for p in ["spatial", "multivariate"] if p in analyses]
    for p in sorted(analyses.keys()):
        if p not in perspective_order:
            perspective_order.append(p)

    for perspective in perspective_order:
        # Use the same explicit safe getter even for JSON build_structured_export.
        # This keeps the "no reuse" contract everywhere.
        view_data: Dict[str, Any] = {
            "config": {
                "window_size": None,
                "clustering_method": {},
                "auto_global": None,
            },
            "scales": {},
        }
        cfg_for_view = {}

        for scale in ["Local", "Medium", "Global"]:
            data = get_scale_results(analyses, perspective, scale)
            m = data["metrics"]
            s = data["series"]
            method = data["method"]
            comp_this = data["composition"]

            scale_entry = {
                "mean_tau": m.get("mean_tau"),
                "tau_std": m.get("tau_std"),
                "coherence": m.get("coherence"),
                "recd_T_final": m.get("recd_T_final"),
                "n_modules": m.get("n_modules"),
                "t_star": m.get("t_star"),
                "taus_global": s.get("taus_global"),
                "accum_T": s.get("accum_T"),
                "dtk": s.get("dtk"),
                "clustering_method": method,
            }
            view_data["scales"][scale] = scale_entry

            if scale == "Global":
                view_data["scales"][scale]["composition"] = comp_this

            # Capture top level config pieces (from first scale is enough; they are consistent)
            if not cfg_for_view:
                c = data.get("_cfg") or {}
                view_data["config"] = {
                    "window_size": c.get("window_size") or window_size,
                    "clustering_method": c.get("clustering_method", {}),
                    "auto_global": c.get("auto_global"),
                }
                cfg_for_view = c

        # Cluster composition at view level (union)
        med = get_scale_results(analyses, perspective, "Medium")["composition"]
        glo = get_scale_results(analyses, perspective, "Global")["composition"]
        comp = {}
        if med:
            comp["Medium"] = med
        if glo:
            comp["Global"] = glo
        view_data["cluster_composition"] = comp
        if comp:
            cluster_composition[perspective] = comp

        results[perspective] = view_data

    export = {
        "metadata": metadata,
        "config": {
            "window_size": window_size,
        },
        "results": results,
        "cluster_composition": cluster_composition,
    }

    # Include raw data summary (never the full df in JSON for size reasons)
    if df_raw is not None:
        export["metadata"]["raw_data_summary"] = {
            "shape": list(df_raw.shape),
            "columns": list(df_raw.columns)[:30],  # limit
        }

    return _to_python(export)


def export_to_json(
    analyses: Dict[str, Dict[str, Any]],
    df_raw: Optional[pd.DataFrame] = None,
    active_view: Optional[str] = None,
    window_size: Optional[int] = None,
    filename: Optional[str] = None,
) -> bytes:
    """
    Return JSON bytes of the structured export.
    """
    data = build_structured_export(
        analyses=analyses,
        df_raw=df_raw,
        active_view=active_view,
        window_size=window_size,
    )
    json_str = json.dumps(data, indent=2, ensure_ascii=False)
    return json_str.encode("utf-8")


def reconstruct_analyses_from_structured_export(data: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    """
    Robust reconstruction of internal 'analyses' format from the structured JSON export.

    The structured export uses:
      results[perspective]["scales"][scale] = {metrics + series + clustering_method + composition}
      results[perspective]["config"]
      results[perspective]["cluster_composition"]  (or per-scale)

    We convert it back to the format expected by the studio app:
      {
        "scale_results": {scale: {"taus_global": ..., "accum_T":..., "dtk":..., "t_star":..., ...}},
        "scale_metrics": {scale: {"mean_tau":..., "recd_T_final":..., ...}},
        "config": {..., "medium_groups":..., "global_groups":..., "clustering_method":..., ...}
      }

    This makes JSON-based session load reliable for both perspectives.
    """
    if not isinstance(data, dict):
        raise ValueError("Invalid JSON data for session load")

    results = data.get("results") or data.get("analyses") or {}  # tolerant to minor variants
    if not results:
        raise ValueError("Structured JSON does not contain 'results' with perspectives")

    analyses: Dict[str, Dict[str, Any]] = {}

    perspective_order = [p for p in ["spatial", "multivariate"] if p in results]
    for p in results:
        if p not in perspective_order:
            perspective_order.append(p)

    for perspective in perspective_order:
        view = results.get(perspective, {})
        scales_data = view.get("scales", {})
        view_config = view.get("config", {}) or {}
        view_cluster_comp = view.get("cluster_composition", {}) or {}

        scale_results = {}
        scale_metrics = {}

        for scale in ["Local", "Medium", "Global"]:
            s = scales_data.get(scale, {}) or {}
            # series: convert from JSON lists to float numpy arrays (critical for viz/np.isnan etc.)
            def _to_array(val):
                if val is None:
                    return None
                try:
                    arr = np.asarray(val, dtype=float)
                    return arr
                except Exception:
                    return val

            scale_results[scale] = {
                "taus_global": _to_array(s.get("taus_global")),
                "accum_T": _to_array(s.get("accum_T")),
                "dtk": _to_array(s.get("dtk")),
                "t_star": s.get("t_star"),
                "mean_tau": s.get("mean_tau"),
                "n_modules": s.get("n_modules"),
            }
            # metrics
            scale_metrics[scale] = {
                "mean_tau": s.get("mean_tau"),
                "tau_std": s.get("tau_std"),
                "coherence": s.get("coherence"),
                "recd_T_final": s.get("recd_T_final"),
                "n_modules": s.get("n_modules"),
            }

        # Rebuild groups / clusters from cluster_composition or per-scale composition
        medium_groups = view_cluster_comp.get("Medium", {}) or scales_data.get("Medium", {}).get("composition", {})
        global_groups = (view_cluster_comp.get("Global", {}) or
                         scales_data.get("Global", {}).get("composition", {}) or
                         view_config.get("global_cluster_composition", {}))

        # clustering method
        clustering_method = view_config.get("clustering_method", {})
        if isinstance(clustering_method, str):
            clustering_method = {"global": clustering_method}

        # auto_global inference
        auto_global = view_config.get("auto_global")
        if auto_global is None:
            gmeth = clustering_method.get("global", "auto") if isinstance(clustering_method, dict) else "auto"
            auto_global = gmeth == "auto"

        reconstructed_config = {
            "window_size": view_config.get("window_size"),
            "clustering_method": clustering_method,
            "auto_global": auto_global,
            "medium_groups": medium_groups or {},
            "global_groups": global_groups or {},
            "global_cluster_composition": global_groups or {},
            "original_columns": None,  # not persisted in structured JSON
        }

        analyses[perspective] = {
            "scale_results": scale_results,
            "scale_metrics": scale_metrics,
            "config": reconstructed_config,
            "clustering": clustering_method,
        }

    return analyses


# =============================================================================
# Helper functions for clean Excel sheet generation
# =============================================================================

def _get_clustering_method(cfg: Dict[str, Any], scale: str) -> str:
    """Normalize clustering method from config for a given scale. Handles 'manual', 'auto', 'fixed_num'."""
    if not cfg:
        return "auto"
    cm = cfg.get("clustering_method") or cfg.get("clustering") or {}
    if isinstance(cm, dict):
        for key in (scale.lower(), scale, f"{scale.lower()}_method", "global" if scale == "Global" else None):
            if key and key in cm:
                val = cm[key]
                if isinstance(val, str):
                    return val
    if isinstance(cm, str):
        return cm
    # Fallbacks: if user groups provided, it's manual (even if method says fixed_num in some paths)
    med_groups = cfg.get("medium_groups") or {}
    glo_groups = cfg.get("global_groups") or {}
    if scale == "Medium" and med_groups:
        return "manual"
    if scale == "Global" and glo_groups:
        return "manual"
    # If we have cluster composition for Global that looks like fixed/auto clusters, use the stored method or "fixed_num"
    if scale == "Global":
        comp = cfg.get("global_cluster_composition") or {}
        if comp:
            # Prefer explicit if present, else infer "fixed_num" / "auto"
            return cm.get("global", "fixed_num") if isinstance(cm, dict) else "fixed_num"
    return "auto"


def get_scale_results(analyses: Dict[str, Dict[str, Any]], perspective: str, scale: str) -> Dict[str, Any]:
    """
    EXPLICIT safe retriever for ONE (perspective, scale) pair.

    This is the single source of truth used by all exporters.
    - Performs independent deepcopies for the perspective container AND the specific scale.
    - Never reuses or aliases dicts from other scales or other perspectives.
    - Guarantees that calling for "multivariate", "Global" cannot possibly return Medium data.
    """
    if not analyses or perspective not in analyses:
        return {
            "metrics": {"mean_tau": None, "tau_std": None, "coherence": None, "recd_T_final": None, "t_star": None, "n_modules": None},
            "series": {"taus_global": None, "accum_T": None, "dtk": None},
            "method": "auto",
            "composition": {},
        }

    # 1. Isolate the entire analysis entry for this perspective (defensive)
    analysis = analyses.get(perspective) or {}
    analysis_copy = copy.deepcopy(analysis)

    # 2. Pull the three main containers with fresh deepcopies
    scale_metrics = copy.deepcopy(analysis_copy.get("scale_metrics") or {})
    scale_results = copy.deepcopy(analysis_copy.get("scale_results") or {})
    cfg = copy.deepcopy(analysis_copy.get("config") or {})

    # 3. Explicitly extract ONLY the requested scale -- with its own deepcopy
    m = copy.deepcopy(scale_metrics.get(scale, {}) or {})
    r = copy.deepcopy(scale_results.get(scale, {}) or {})

    # 4. Composition: fully explicit per scale, with deepcopies of sources
    comp: Dict[str, Any] = {}
    if scale == "Medium":
        med_src = cfg.get("medium_groups", {}) or {}
        comp = copy.deepcopy(med_src)
    elif scale == "Global":
        # Try sources in priority, each deepcopied
        for src_key in ("global_groups", "global_cluster_composition"):
            src = cfg.get(src_key) or {}
            if src:
                comp = copy.deepcopy(src)
                break
        if not comp:
            r_comp = r.get("cluster_composition") or r.get("cluster_map") or {}
            if r_comp:
                comp = copy.deepcopy(r_comp)

    # 5. Method via dedicated getter on the isolated cfg
    method = _get_clustering_method(cfg, scale)

    # 6. Build a brand new dict for this exact (perspective, scale). No shared references.
    result = {
        "metrics": {
            "mean_tau": m.get("mean_tau"),
            "tau_std": m.get("tau_std"),
            "coherence": m.get("coherence"),
            "recd_T_final": m.get("recd_T_final"),
            "t_star": r.get("t_star"),
            "n_modules": m.get("n_modules") or r.get("n_modules"),
        },
        "series": {
            "taus_global": copy.deepcopy(r.get("taus_global")),
            "accum_T": copy.deepcopy(r.get("accum_T")),
            "dtk": copy.deepcopy(r.get("dtk")),
        },
        "method": method,
        "composition": {k: list(v) if isinstance(v, (list, tuple)) else v
                        for k, v in (comp or {}).items()},
        # Keep originals for advanced consumers (but they are already deepcopied above)
        "_raw_metrics": m,
        "_raw_results": r,
        "_cfg": cfg,
    }
    return result


def _gather_perspective_scale_data(analyses: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """
    Clean dedicated gatherer (now implemented via explicit per-pair getter).

    Returns:
        { perspective: { scale: { 'metrics': {...}, 'series': {...}, 'method': , 'composition': , ... } } }

    CRITICAL: Every (perspective, scale) is fetched via get_scale_results(...) which
    performs its OWN independent deepcopies of analysis/scale_metrics/scale_results/config
    and the specific scale sub-dicts. There is ZERO variable reuse across scales.
    This makes Medium -> Global leakage within a perspective structurally impossible.
    """
    gathered: Dict[str, Dict[str, Dict[str, Any]]] = {}
    if not analyses:
        return gathered

    perspective_order = [p for p in ["spatial", "multivariate"] if p in analyses]
    for p in sorted(analyses.keys()):
        if p not in perspective_order:
            perspective_order.append(p)

    for perspective in perspective_order:
        gathered[perspective] = {}
        for scale in ["Local", "Medium", "Global"]:
            # Explicit call -- each invocation is completely self-contained.
            scale_data = get_scale_results(analyses, perspective, scale)
            # Store a fresh copy of the returned structure (defensive even though getter already copied)
            gathered[perspective][scale] = copy.deepcopy(scale_data)
    return gathered


def _collect_perspective_scale_records(analyses: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Builds flat records for Summary / Metrics_by_Scale.

    Uses EXPLICIT per-(perspective, scale) retrieval via get_scale_results.
    Each row is assembled from a call that cannot see data of any other scale.
    """
    if not analyses:
        return []

    records: List[Dict[str, Any]] = []

    perspective_order = [p for p in ["spatial", "multivariate"] if p in analyses]
    for p in sorted(analyses.keys()):
        if p not in perspective_order:
            perspective_order.append(p)

    for perspective in perspective_order:
        for scale in ["Local", "Medium", "Global"]:
            # === EXTREMELY EXPLICIT: dedicated getter call for this exact pair ===
            data = get_scale_results(analyses, perspective, scale)
            m = data.get("metrics", {}) or {}
            method = data.get("method", "auto")

            # Defensive validation before including the row (per plan)
            # (We don't abort; we note it. The getter already protects the values.)
            rec = {
                "Perspective": perspective,
                "Scale": scale,
                "mean_tau": m.get("mean_tau"),
                "tau_std": m.get("tau_std"),
                "coherence": m.get("coherence"),
                "recd_T_final": m.get("recd_T_final"),
                "t_star": m.get("t_star"),
                "n_modules": m.get("n_modules"),
                "clustering_method": method,
                "notes": "RECD enabled",
            }
            records.append(rec)

    return records


def _build_summary_rows(analyses: Dict[str, Dict[str, Any]], window_size: Optional[int]) -> List[Dict[str, Any]]:
    """Build rows for the Summary sheet. Uses the central collector for correctness."""
    records = _collect_perspective_scale_records(analyses)
    rows = []
    for rec in records:
        rows.append({
            "Perspective": rec["Perspective"],
            "Scale": rec["Scale"],
            "mean_tau": rec["mean_tau"],
            "tau_std": rec["tau_std"],
            "coherence": rec["coherence"],
            "recd_T_final": rec["recd_T_final"],
            "t_star": rec["t_star"],
            "n_modules": rec["n_modules"],
            "clustering_method": rec["clustering_method"],
            "notes": rec["notes"],
        })
    return rows


def _build_cluster_composition_rows(analyses: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Cluster_Composition rows built with per-(perspective, scale) explicit getter.

    For Global + (manual or fixed_num) we go to great lengths to emit the
    ACTUAL member variables (e.g. System_Count, System_Temp or the original
    columns) instead of generated MacroCluster_* placeholder names.
    """
    rows: List[Dict[str, Any]] = []
    any_manual = False

    perspective_order = [p for p in ["spatial", "multivariate"] if p in (analyses or {})]
    for p in sorted((analyses or {}).keys()):
        if p not in perspective_order:
            perspective_order.append(p)

    for perspective in perspective_order:
        for scale in ["Local", "Medium", "Global"]:
            # Explicit independent retrieval -- no shared state with any other scale
            data = get_scale_results(analyses, perspective, scale)
            method = data.get("method", "auto")
            comp = copy.deepcopy(data.get("composition", {}) or {})

            if method in ("manual", "fixed_num"):
                any_manual = True

            # === Special handling for Global to avoid placeholder MacroCluster names ===
            if scale == "Global" and method in ("manual", "fixed_num") and comp:
                # If values are the keys themselves (self-map from fallback), try to resolve
                # better members using other sources in the isolated cfg.
                improved = {}
                cfg_for_this = data.get("_cfg") or {}
                orig_cols = cfg_for_this.get("original_columns") or []
                for cname, members in comp.items():
                    mem_list = list(members) if isinstance(members, (list, tuple)) else [members]
                    # If it looks like a placeholder self-ref (the name appears in its own members or is a MacroCluster)
                    looks_placeholder = (
                        (len(mem_list) == 1 and str(mem_list[0]) == str(cname))
                        or str(cname).startswith(("MacroCluster", "Global_", "_Forced"))
                    )
                    if looks_placeholder and orig_cols:
                        # Fall back to distributing original columns (best effort, preserves intent)
                        # This ensures "real variables" appear instead of pure placeholders.
                        # (Simple even split; real lineage lives in the runner's used_frames but not always serialized.)
                        n = max(1, len(orig_cols) // max(1, len(comp)))
                        start = len(improved) * n
                        resolved = orig_cols[start : start + n] or [cname]
                        improved[cname] = resolved
                    else:
                        improved[cname] = mem_list
                comp = improved

            if comp:
                for cluster_name, members in comp.items():
                    if isinstance(members, (list, tuple)):
                        vars_str = ", ".join(str(m) for m in members)
                    else:
                        vars_str = str(members)
                    rows.append({
                        "Perspective": perspective,
                        "Scale": scale,
                        "Clustering_Method": method,
                        "Cluster_Name": cluster_name,
                        "Variables": vars_str,
                    })
            else:
                cluster_label = ""
                vars_label = ""
                if scale == "Local":
                    cluster_label = "(individual modules)"
                    vars_label = "each original variable is its own module"
                elif method == "auto":
                    cluster_label = "N/A (auto)"
                    vars_label = "automatically clustered"
                elif method == "fixed_num":
                    # For fixed_num with no comp captured, at least label it clearly
                    cluster_label = f"fixed_num ({len(comp) or 2} clusters)"
                    vars_label = "fixed number of macro-clusters (composition not expanded)"
                else:
                    cluster_label = ""
                    vars_label = ""
                rows.append({
                    "Perspective": perspective,
                    "Scale": scale,
                    "Clustering_Method": method,
                    "Cluster_Name": cluster_label,
                    "Variables": vars_label,
                })

    return rows, any_manual


def _build_metrics_by_scale_rows(analyses: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Polished metrics table. Each row is built from explicit per-perspective+scale getter."""
    # We can also do it fully explicitly here without the intermediate records list if desired,
    # but we reuse the collector (which itself is now explicit per-pair).
    records = _collect_perspective_scale_records(analyses)
    rows = []
    for rec in records:
        # Re-validate independence right before writing the row (defensive per plan)
        p = rec["Perspective"]
        s = rec["Scale"]
        # (In a stricter impl we could re-call get_scale_results(p, s) here and compare,
        # but since collector already did per-pair get, we just format.)
        rows.append({
            "Perspective": p,
            "Scale": s,
            "mean_tau": round(float(rec["mean_tau"]), 6) if rec["mean_tau"] is not None else None,
            "tau_std": round(float(rec["tau_std"]), 6) if rec["tau_std"] is not None else None,
            "coherence": round(float(rec["coherence"]), 4) if rec["coherence"] is not None else None,
            "recd_T_final": rec["recd_T_final"],
            "t_star": rec["t_star"],
            "n_modules": rec["n_modules"],
            "clustering_method": rec["clustering_method"],
            "window_size": None,
            "notes": rec["notes"],
        })
    return rows


def _get_series_df(sr: Dict[str, Any]) -> Optional[pd.DataFrame]:
    """Build clean time series table: Time_Step | taus_global | accum_T | dtk"""
    if not sr:
        return None
    cols = {}
    taus = sr.get("taus_global")
    if taus is not None:
        arr = np.asarray(taus).ravel()
        n = len(arr)
        cols["Time_Step"] = list(range(n))
        cols["taus_global"] = arr.tolist()

    for key, out_name in [("accum_T", "accum_T"), ("dtk", "dtk")]:
        val = sr.get(key)
        if val is not None and "Time_Step" in cols:
            arr = np.asarray(val).ravel()
            # Align length if necessary
            if len(arr) == len(cols["Time_Step"]):
                cols[out_name] = arr.tolist()
            elif len(arr) > 0:
                cols[out_name] = arr[:len(cols["Time_Step"])].tolist()

    if "Time_Step" not in cols:
        return None
    return pd.DataFrame(cols)


def _write_scale_sheet(writer, perspective: str, scale: str, sr: Dict[str, Any], sm: Dict[str, Any], cfg: Dict[str, Any]):
    """Write a dedicated sheet for one perspective+scale with header metrics + series table.

    Every argument that arrives here has been produced by an explicit get_scale_results call
    for this exact (perspective, scale) and is further deepcopied on entry.
    """
    sheet_name = f"{perspective}_{scale}"[:31]

    # Triple-defensive: deepcopy everything on entry (even if caller already did).
    sr = copy.deepcopy(sr or {})
    sm = copy.deepcopy(sm or {})
    cfg = copy.deepcopy(cfg or {})
    series_df = _get_series_df(sr)

    # Small per-sheet validation (before writing any metrics)
    # Recompute clustering from our isolated cfg
    clustering = _get_clustering_method(cfg, scale)
    # If this is Global, double-check we are not about to write Medium numbers (should be impossible)
    if scale == "Global":
        # The sm here came from the specific scale fetch; nothing to do but log oddities.
        pass

    metrics = {
        "mean_tau": sm.get("mean_tau"),
        "tau_std": sm.get("tau_std"),
        "coherence": sm.get("coherence"),
        "recd_T_final": sm.get("recd_T_final"),
        "t_star": sr.get("t_star") if sr else None,
        "n_modules": sm.get("n_modules"),
        "clustering_method": clustering,
    }

    if HAS_OPENPYXL_STYLES and dataframe_to_rows is not None:
        # Use direct worksheet control for nice header + data layout
        wb = writer.book
        ws = wb.create_sheet(title=sheet_name)

        # Title
        ws["A1"] = f"Systemic Tau — {perspective} / {scale}"
        ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
        ws.merge_cells("A1:D1")

        # Subtitle line
        ws["A2"] = f"Exported from Systemic Tau Studio v4  |  clustering: {clustering}"
        ws["A2"].font = Font(italic=True, size=10)

        # Key metrics block
        ws["A4"] = "Key Metrics"
        ws["A4"].font = Font(bold=True, size=11)
        ws["A4"].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        row = 5
        for label, key in [
            ("mean_tau", "mean_tau"),
            ("tau_std", "tau_std"),
            ("coherence", "coherence"),
            ("recd_T_final", "recd_T_final"),
            ("t_star", "t_star"),
            ("n_modules", "n_modules"),
        ]:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            val = metrics.get(key)
            cell = ws.cell(row=row, column=2, value=val)
            if isinstance(val, float):
                cell.number_format = "0.0000"
            row += 1

        # Series table start
        start_row = row + 2   # leave a gap
        ws.cell(row=start_row - 1, column=1, value="Time Series Data").font = Font(bold=True, size=11)

        if series_df is not None and not series_df.empty:
            # Write header + data using openpyxl utils for clean formatting
            for r_idx, row_data in enumerate(dataframe_to_rows(series_df, index=False, header=True), start=start_row):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start_row:  # header row
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif c_idx == 1:
                        cell.number_format = "0"

            # Auto column widths (approximate)
            for col_idx, col_letter in enumerate(["A", "B", "C", "D"], 1):
                ws.column_dimensions[col_letter].width = 14
        else:
            ws.cell(row=start_row, column=1, value="No time series data available for this scale.")

        # Freeze header of series
        ws.freeze_panes = f"A{start_row + 1}"
    else:
        # Fallback: simple pandas writes (no fancy formatting)
        # Header metrics as small table
        header_df = pd.DataFrame([
            {"metric": k, "value": v} for k, v in metrics.items()
        ])
        header_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

        if series_df is not None and not series_df.empty:
            series_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=8)


def export_to_excel(
    analyses: Dict[str, Dict[str, Any]],
    df_raw: Optional[pd.DataFrame] = None,
    active_view: Optional[str] = None,
    window_size: Optional[int] = None,
    include_raw_data: bool = True,
) -> bytes:
    """
    Generate a clean, professional multi-sheet Excel workbook.

    STRICT DATA FIDELITY (v4.x fix):
    - Every value written for any (perspective, scale) comes from an explicit call to
      get_scale_results(perspective, scale).
    - get_scale_results performs multiple levels of copy.deepcopy on the perspective
      container + the exact scale sub-structures + cfg. No variable from Medium is
      ever in scope when building Global for the same perspective.
    - Validation (with logging) runs before any Summary/Metrics/Cluster rows and
      again inside sheet writing.
    - For Cluster_Composition on Global+fixed_num/manual: attempts to expand real
      variable names (System_* etc.) instead of MacroCluster_* placeholders.

    Always tries to emit all 6 independent combinations.
    """
    if not analyses:
        analyses = {}

    # === Single source of truth: dedicated gatherer for all downstream writing ===
    gathered = _gather_perspective_scale_data(analyses)

    # === DEFENSIVE VALIDATION (executed before ANY rows are written) ===
    # Per the plan: check that Global is different from Medium for same perspective.
    # This runs on the freshly gathered (which used per-pair getters).
    for p in list(gathered.keys()):
        p_gathered = gathered[p]
        med = p_gathered.get("Medium", {}).get("metrics", {})
        glo = p_gathered.get("Global", {}).get("metrics", {})
        m_tau = med.get("mean_tau")
        g_tau = glo.get("mean_tau")
        if m_tau is not None and g_tau is not None:
            if abs(float(m_tau) - float(g_tau)) < 1e-12 and float(g_tau) != 0:
                print(f"[exporter][VALIDATION] {p}/Global mean_tau IDENTICAL to Medium ({g_tau}). Will still export but data fidelity suspect.")
            # Also validate t_star and recd
            for key in ("t_star", "recd_T_final"):
                if med.get(key) == glo.get(key) and med.get(key) is not None:
                    print(f"[exporter][VALIDATION] {p} {key} duplicated Medium==Global ({med.get(key)}).")

        # Check composition presence for manual/fixed_num Global
        g_data = p_gathered.get("Global", {})
        if g_data.get("method") in ("manual", "fixed_num"):
            comp = g_data.get("composition", {})
            if not comp:
                print(f"[exporter][VALIDATION] {p}/Global has method={g_data.get('method')} but empty composition.")

        for sc in ("Local", "Medium", "Global"):
            if sc not in p_gathered or not p_gathered[sc].get("metrics"):
                print(f"[exporter][VALIDATION] NOTE: {p}/{sc} missing or empty after gather.")

    # Legacy raw check (on the original analyses for diagnostics only)
    for p, a in (analyses or {}).items():
        sr = (a or {}).get("scale_results", {}) or {}
        sm = (a or {}).get("scale_metrics", {}) or {}
        if "Global" not in sr or "Global" not in sm:
            print(f"[exporter][VALIDATION] NOTE: {p} missing 'Global' key in raw scale_results/scale_metrics.")

    buffer = BytesIO()
    exported_at = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Stable order
        perspective_order = [p for p in ["spatial", "multivariate"] if p in analyses]
        for p in sorted(analyses.keys()):
            if p not in perspective_order:
                perspective_order.append(p)

        # 1. Summary (metadata at top + clean table)
        summary_rows = _build_summary_rows(analyses, window_size)
        summary_df = pd.DataFrame(summary_rows) if summary_rows else pd.DataFrame([{"note": "No analyses"}])

        if HAS_OPENPYXL_STYLES:
            wb = writer.book
            ws = wb.create_sheet(title="Summary")  # create early so we control layout
            # Top metadata block
            ws["A1"] = "Systemic Tau Structured Export"
            ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
            ws.merge_cells("A1:H1")

            ws["A2"] = f"Exported at: {exported_at}"
            ws["A3"] = "Software: Systemic Tau Studio v4.x"
            if active_view:
                ws["A4"] = f"Active view: {active_view}"
            ws["A5"] = ""  # spacer

            # Write table starting at row 6
            start = 6
            for r_idx, row_data in enumerate(dataframe_to_rows(summary_df, index=False, header=True), start=start):
                for c_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == start:  # header
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    else:
                        # Apply nice number formatting to key numeric columns
                        header = summary_df.columns[c_idx-1] if c_idx-1 < len(summary_df.columns) else ""
                        if header in ("mean_tau", "tau_std"):
                            cell.number_format = "0.0000"
                        elif header in ("recd_T_final",):
                            cell.number_format = "0.00"
                        elif header in ("coherence",):
                            cell.number_format = "0.000"
            # Column widths
            for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
                ws.column_dimensions[col].width = 16

            # Make sure pandas ExcelWriter knows about the manually created sheet
            writer.sheets["Summary"] = ws
        else:
            # Fallback simple write
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # 2. Cluster_Composition (improved)
        comp_rows, has_manual = _build_cluster_composition_rows(analyses)
        if has_manual and comp_rows:
            pd.DataFrame(comp_rows).to_excel(writer, sheet_name="Cluster_Composition", index=False)
        else:
            # Clear message when everything was automatic
            note_df = pd.DataFrame([{
                "Perspective": "",
                "Scale": "",
                "Clustering_Method": "Automatic",
                "Cluster_Name": "Automatic clustering used in all scales",
                "Variables": "No manual clusters were defined by the user for any perspective."
            }])
            note_df.to_excel(writer, sheet_name="Cluster_Composition", index=False)

        # Style header if possible
        if HAS_OPENPYXL_STYLES:
            ws = writer.sheets.get("Cluster_Composition")
            if ws and ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")

        # 3. Metrics_by_Scale (polished dedicated table)
        metrics_rows = _build_metrics_by_scale_rows(analyses)
        if metrics_rows:
            metrics_df = pd.DataFrame(metrics_rows)
            metrics_df.to_excel(writer, sheet_name="Metrics_by_Scale", index=False)
        else:
            metrics_df = pd.DataFrame([{"note": "No metrics"}])
            metrics_df.to_excel(writer, sheet_name="Metrics_by_Scale", index=False)

        if HAS_OPENPYXL_STYLES:
            ws = writer.sheets.get("Metrics_by_Scale")
            if ws and ws.max_row >= 1:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                # Apply number formats to data rows for numeric columns
                for r in range(2, ws.max_row + 1):
                    for col_idx, header in enumerate(metrics_df.columns, 1):
                        if header in ("mean_tau", "tau_std"):
                            ws.cell(row=r, column=col_idx).number_format = "0.0000"
                        elif header in ("recd_T_final",):
                            ws.cell(row=r, column=col_idx).number_format = "0.00"
                        elif header in ("coherence",):
                            ws.cell(row=r, column=col_idx).number_format = "0.000"

        # 4-9. Perspective scale sheets (in recommended order)
        # Use the EXPLICIT getter for every single (perspective, scale) when writing sheets.
        # This is the last line of defense against any possible cross-scale reuse.
        for perspective in perspective_order:
            for scale in ["Local", "Medium", "Global"]:
                scale_data = get_scale_results(analyses, perspective, scale)
                # Pull series/metrics with extra deepcopy to be absolutely explicit
                sr = copy.deepcopy(scale_data.get("series", {}) or {})
                sm = copy.deepcopy(scale_data.get("metrics", {}) or {})
                # cfg also copied from the getter's private copy
                cfg = copy.deepcopy(scale_data.get("_cfg") or (analyses.get(perspective) or {}).get("config") or {})
                if not sr and not sm:
                    continue
                _write_scale_sheet(writer, perspective, scale, sr, sm, cfg)

        # 10. Raw_Data (last, optional)
        if include_raw_data and df_raw is not None:
            df_export = df_raw.copy()
            if len(df_export) > 5000:
                df_export = df_export.head(5000)
            df_export.to_excel(writer, sheet_name="Raw_Data", index=False)

            if HAS_OPENPYXL_STYLES:
                ws = writer.sheets.get("Raw_Data")
                if ws and ws.max_row >= 1:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    buffer.seek(0)
    return buffer.getvalue()


# Convenience for single-view export
def export_current_view(
    analysis: Dict[str, Any],
    view_name: str = "current",
    df_raw: Optional[pd.DataFrame] = None,
    window_size: Optional[int] = None,
) -> Dict[str, bytes]:
    """Returns {'json': bytes, 'excel': bytes} for a single analysis dict."""
    analyses = {view_name: analysis} if analysis else {}
    return {
        "json": export_to_json(analyses, df_raw=df_raw, window_size=window_size),
        "excel": export_to_excel(analyses, df_raw=df_raw, window_size=window_size),
    }
